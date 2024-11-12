import os
import fitz  # PyMuPDF for PDF handling
from PIL import Image, ImageEnhance
import pytesseract
import shutil
import io
import pandas as pd
from datetime import datetime
from pdf2image import convert_from_path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from concurrent.futures import ThreadPoolExecutor, as_completed
from transformers import pipeline  # For text summarization
from threading import Lock

# Initialize summarizer
summarizer = pipeline("summarization", model="facebook/bart-large-cnn")

file_lock = Lock()

def load_complaint_keywords(file_path="./keywords/complaint_keywords.txt"):
    """Load complaint keywords from a text file, each keyword on a new line."""
    keywords = []
    if os.path.exists(file_path):
        with open(file_path, "r") as f:
            keywords = [line.strip().lower() for line in f if line.strip()]
    else:
        print(f"[WARNING] Keywords file '{file_path}' not found. Using default keywords.")
    return keywords or [
        'complaint', 'grievance', 'mad', 'angry', 'upset', 'legal', 'action', 'legal action', 'lawyer', 'frustration', 'attorney',
        'department of insurance', 'doi', 'lawyer', 'regulatory agency', 'lawsuit', 'hire', 'deadline', 'action',
        'unauthorized', 'inappropriate', 'theft', 'forgery', 'fraud', 'media', 'better business bureau',
        'subpoena', 'attorney letterhead'
    ]

# Load complaint keywords from file
complaint_keywords = load_complaint_keywords()

def extract_text_from_pdf(pdf_file):
    """Extract text from each page of a PDF file, with OCR fallback for image-based pages."""
    doc = fitz.open(pdf_file)
    text = ""
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        page_text = page.get_text()
        
        if not page_text.strip():  # If page text is empty, perform OCR
            pixmap = page.get_pixmap()
            img = Image.open(io.BytesIO(pixmap.tobytes("png")))
            page_text = pytesseract.image_to_string(img)
        
        text += page_text + "\n\n"
    doc.close()
    return text

def ocr_image_file(image_file):
    """Extract text from a single image or multi-page TIF file using OCR."""
    text = ""
    img = Image.open(image_file)
    
    # Iterate through pages if TIF is multi-page
    if hasattr(img, "n_frames"):
        for page in range(img.n_frames):
            img.seek(page)
            page_text = pytesseract.image_to_string(img)
            text += page_text + "\n\n"
    else:
        text += pytesseract.image_to_string(img)
    return text

def process_file(file_path):
    """Process a file (PDF or TIF) and return the extracted text."""
    file_ext = os.path.splitext(file_path)[1].lower()
    text = ""
    
    if file_ext == '.pdf':
        text = extract_text_from_pdf(file_path)
        if not text.strip():  # If no text found, treat as scanned PDF and use OCR
            images = convert_from_path(file_path)
            for img in images:
                text += pytesseract.image_to_string(img) + "\n\n"
    elif file_ext in ['.tif', '.tiff']:
        text = ocr_image_file(file_path)
    else:
        print(f"[ERROR] Unsupported file format: {file_ext}")
    
    return text

def classify_text(text):
    """Classify text based on keywords for 'complaint' or 'Appeal'."""
    text_lower = text.lower()
    return 'Complaint' if any(keyword in text_lower for keyword in complaint_keywords) else 'Appeal'

def summarize_text(text):
    """Generate a summary of the text if it is long enough, with error handling."""
    try:
        if len(text) > 1024:  # Summarize in chunks if text is large
            summaries = []
            for i in range(0, len(text), 1024):
                chunk = text[i:i + 1024]
                summaries.append(summarizer(chunk, max_length=60, min_length=25, do_sample=False)[0]["summary_text"])
            return " ".join(summaries)
        else:
            return summarizer(text, max_length=60, min_length=25, do_sample=False)[0]["summary_text"]
    except Exception as e:
        print(f"Error during summarization: {e}")
        return "Summary not available"

def style_excel(filename):
    """Style the generated Excel report for better readability."""
    workbook = load_workbook(filename)
    sheet = workbook.active
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 50  # Adjust column width for summary
    workbook.save(filename)
    workbook.close()

def process_single_file(idx, file_name, input_folder, extracted_folder, complaint_folder, appeal_folder, unable_to_detect_folder):
    file_path = os.path.join(input_folder, file_name)
    
    with file_lock:
        if any(os.path.exists(os.path.join(folder, file_name)) for folder in [complaint_folder, appeal_folder, unable_to_detect_folder]):
            print(f"[INFO] Skipping already processed file: {file_name}")
            return None

    text = process_file(file_path)
    classification = classify_text(text)
    summary = summarize_text(text) if text else "No content to summarize"
    
    # Save extracted text in the respective folder based on classification
    output_folder = os.path.join(extracted_folder, classification)
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, f"{os.path.splitext(file_name)[0]}_output.txt")
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(text)

    processed_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    result = [idx, file_name, processed_time, classification, summary]

    # Determine the destination folder for archiving
    destination_folder = complaint_folder if classification == 'Complaint' else appeal_folder
    archive_path = os.path.join(destination_folder, file_name)
    shutil.move(file_path, archive_path)

    print(f"[INFO] Processed {file_name}, classified as {classification} and saved output to {output_file}")
    print(f"[INFO] Moved {file_name} to {destination_folder}")
    
    return result

def main():
    input_folder = 'input'
    extracted_folder = 'extracted_text'
    archive_folder = 'archive'
    reports_folder = 'classification_reports'

    # Create required directories if they don't exist
    os.makedirs(input_folder, exist_ok=True)
    os.makedirs(extracted_folder, exist_ok=True)
    os.makedirs(archive_folder, exist_ok=True)
    os.makedirs(reports_folder, exist_ok=True)

    complaint_folder = os.path.join(archive_folder, 'Complaint')
    appeal_folder = os.path.join(archive_folder, 'Appeal')
    unable_to_detect_folder = os.path.join(archive_folder, 'Manual')

    os.makedirs(complaint_folder, exist_ok=True)
    os.makedirs(appeal_folder, exist_ok=True)
    os.makedirs(unable_to_detect_folder, exist_ok=True)

    results = []

    # Define unique Excel filename based on the current date and time
    report_time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    results_file = os.path.join(reports_folder, f'classification_results_{report_time}.xlsx')

    with ThreadPoolExecutor(max_workers=4) as executor:
        future_to_file = {
            executor.submit(
                process_single_file,
                idx, file_name,
                input_folder,
                extracted_folder,
                complaint_folder,
                appeal_folder,
                unable_to_detect_folder
            ): file_name
            for idx, file_name in enumerate(os.listdir(input_folder), start=1)
        }

        for future in as_completed(future_to_file):
            result = future.result()
            if result:
                results.append(result)

    # Save results to the new timestamped Excel file
    combined_df = pd.DataFrame(results, columns=['Serial No', 'File Name', 'File Processed Date and Time', 'Classification', 'Summary'])
    combined_df.to_excel(results_file, index=False, engine='openpyxl')
    
    # Style the new Excel file
    style_excel(results_file)
    
    print(f"[INFO] Classification report saved to {results_file}")

if __name__ == "__main__":
    main()
